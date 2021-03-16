'''
Created on Feb. 9, 2020

@author: cefect

cost benefit calculations
'''

#==========================================================================
# logger setup-----------------------
#==========================================================================
import logging, configparser, datetime, copy, shutil



#==============================================================================
# imports------------
#==============================================================================
import os
import numpy as np
import pandas as pd




#==============================================================================
# dependency check
#==============================================================================
# Let users know if they're missing any of our hard dependencies
hard_dependencies = ("openpyxl",)
missing_dependencies = []

for dependency in hard_dependencies:
    try:
        __import__(dependency)
    except ImportError as e:
        missing_dependencies.append("{0}: {1}".format(dependency, str(e)))

if missing_dependencies:
    raise ImportError(
        "Unable to import required dependencies:\n" + "\n".join(missing_dependencies)
    )
    
del hard_dependencies, dependency, missing_dependencies


import openpyxl

#===============================================================================
# customs
#===============================================================================
from hlpr.exceptions import QError as Error
from hlpr.basic import view
 
from results.riskPlot import RiskPlotr

#==============================================================================
# functions-------------------
#==============================================================================
class CbaWrkr(RiskPlotr):
 
    
    cba_xls =None
    
    #===========================================================================
    # expectations from parameter file
    #===========================================================================
    exp_pars_md = {
        'results_fps':{
             'r_ttl':{'ext':('.csv',)},
             }
        }
    
    exp_pars_op={
        'results_fps':{
             'cba_xls':{'ext':('.xlsx',)},
             }
        }
    
    def __init__(self,
                figsize=(10,6),
                 **kwargs):
        
        super().__init__(figsize=figsize, **kwargs)
        
        self.dtag_d={**self.dtag_d,**{
            'r_ttl':{'index_col':None}}}
        
        #=======================================================================
        # paramters directory
        #=======================================================================
        self.pars_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), '_pars')
        
        #=======================================================================
        # get the temlpate file
        #=======================================================================
        self.template_fp = os.path.join(self.pars_dir, 'cf_bca_template_01.xlsx')

        assert os.path.exists(self.template_fp), 'passed template_fp is bad: \'%s\''%self.template_fp
        
        self.logger.debug('%s.__init__ w/ feedback \'%s\''%(
            self.__class__.__name__, type(self.feedback).__name__))
        
    def prep_model(self):

        
        self.set_ttl() #load and prep the total results

        
        return 
        
    def copy_template(self, #copy the cba template worksheet
                          template_fp = None,
                          logger=None,
                          ):
        
        #=======================================================================
        # defaults
        #=======================================================================
        if logger is None: logger=self.logger
        log=logger.getChild('copy_template')
        
        if template_fp is None: template_fp=self.template_fp
        
        #=======================================================================
        # prechecks
        #=======================================================================
        assert os.path.exists(template_fp), 'passed template_fp is bad: \'%s\''%template_fp
        
        #=======================================================================
        # load
        #=======================================================================
        wbook = openpyxl.load_workbook(template_fp)
        wsheet = wbook['smry']
        log.info('loaded worksheet from: \n    %s'%template_fp)
        #=======================================================================
        # add data----
        #=======================================================================
        
        #=======================================================================
        # scenario description
        #=======================================================================
        meta_d = dict() #collect for reporting
        #loop through select rows, check, and set the new value
        for evalA, nvalB in {
            'name':self.name,
            'control_filename':os.path.basename(self.cf_fp),
            'ead_option':self.ead_tot,
            'timestamp':self.today_str,
            'base_year':int(datetime.datetime.now().strftime('%Y')),
            }.items():
            
            if nvalB is None or nvalB=='':
                log.warning('failed to get a value for \'%s\'... skippin'%evalA)
            #===================================================================
            # #find first location with this expectstion
            #===================================================================
            cellA = None
            for row in wsheet.iter_rows(max_col=1, max_row=30, min_row=5):
                for c in row:
                    if c.value==evalA:
                        cellA = c
                        break #stop heere and  use this cell
            
            #check fail
            if cellA is None:
                log.warning('unable to locate \'%s\' on template... skipping'%evalA)
                continue #skip 
                
            log.debug('found \'%s\' at %s'%(evalA, cellA.coordinate))
            
            #===================================================================
            # set the value
            #===================================================================
            #get value cell
            cellB = wsheet.cell(row=cellA.row, column=2)
            
            #set the new value
            cellB.value=nvalB
            
            #report
            log.debug('set %s=%s'%(cellB.coordinate, nvalB))
            meta_d[cellB.coordinate] = cellB.value
            
        log.info('updated template \'smry\' tab w/\n      %s'%meta_d)
        
    
        #=======================================================================
        # wrap
        #=======================================================================
        self.wbook = wbook #set for saving
        return self.wbook
    
    def write_wbook(self, #helper to write the openpyxl workbo0ok to file
                       wbook=None,
                       ofp=None,
                       logger=None,
                       upd_cf=None,
                       ):
        
        #=======================================================================
        # defaults
        #=======================================================================
        if wbook is None: wbook=self.wbook
        if ofp is None: ofp = os.path.join(self.out_dir, 'cba_%s_%s.xlsx'%(self.name, self.tag))
        if logger is None: logger=self.logger
        if upd_cf is  None: upd_cf=self.upd_cf
        log=logger.getChild('write_wbook')
        
        #=======================================================================
        # write
        #=======================================================================
        wbook.save(ofp)
        log.info('wrote workbook to file \n    %s'%(ofp))
        
        #=======================================================================
        # update control file
        #=======================================================================
        if upd_cf:
            self.set_cf_pars(
                {'results_fps': ({'cba_xls':ofp}, '#%s set cba_xls on %s'%(
                    self.__class__.__name__, self.today_str))}
                )

        
        return ofp
    
    def plot_cba(self, #generate a plot of costs and benefits
                 data_fp=None,
                 
                 presentVal=True, #whether to adjust values to present
                 
                 logger=None,
                 title=None,
                 
                 #plot styles
                 impactFmtFunc = None, #formattting the y axis
                 style_lib = { #styles to apply to each line
                     'Grand Total Costs':{
                            'color':'red', 
                            'marker':'x', 'markersize':3,
                                },
                    'Grand Total Benefits':{
                            'color':'green',
                            'marker':'x', 'markersize':3,
                                },
                    'hatch':{
                            'alpha':0.5,
                            'hatch':None,
                            }
                     }
                 ):
        
        #=======================================================================
        # defaults
        #=======================================================================
        if data_fp is None: data_fp = self.cba_xls
        assert os.path.exists(data_fp), 'passed bad filepath for cba_xls: %s'%data_fp
        
        if logger is None: logger=self.logger
        log=logger.getChild('plot_cba')
        
        if title is None: title = '%s Cost-Benefit Curves for %s'%(self.tag, self.name)
        if impactFmtFunc is None: impactFmtFunc=self.impactFmtFunc
        
        plt, matplotlib = self.plt, self.matplotlib
        #=======================================================================
        # get data------
        #=======================================================================
        #=======================================================================
        # load  
        #=======================================================================
        df_raw = pd.read_excel(data_fp, sheet_name='data' ,engine='openpyxl')
        
        log.info('loaded %s from %s'%(str(df_raw.shape), data_fp))
        
        #=======================================================================
        # check
        #=======================================================================
        assert 'plot' in df_raw.columns
        
        #=======================================================================
        # clean DATA
        #=======================================================================
        df = df_raw.dropna(axis=0, how='all').copy() #drop empty rows
        
        #fix plot bool
        df.loc[:, 'plot'] = df['plot'].fillna(0).astype(bool)
        
        #get just rows w/ plot=Truee
        df = df.loc[df['plot'], :].drop('plot',axis=1)
        
        #fix index
        df = df.set_index(df.columns[0], drop=True)
        df.index.name = None
        
        #check data
        assert not df.isna().any().any(), 'got some nulls... make sure all the data is complete and the spreadsheet saved'
        
        #drop any remaining unamed columns
        boolcol = df.columns.str.contains('Unnamed:').fillna(False).values.astype(bool)
        
        """
        view(df)
        """
        
        
        df = df.loc[:, np.invert(boolcol)]
        
        log.debug('cleaned cba data to %s'%(str(df.shape)))
        
        #minmum length checks
        assert len(df.index)>=2
        assert len(df.columns)>=2
        assert df.notna().all().all(), 'got some nulls... make sure all the data is complete'
        #=======================================================================
        # get cumulatives
        #=======================================================================
        dfc = df.cumsum(axis=1) #convert to cumulatives
        
        #=======================================================================
        # get label data
        #=======================================================================
        try:
            calc_d = self._get_cba_calcs(data_fp=data_fp, impactFmtFunc=impactFmtFunc, logger=log)
        except Exception as e:
            log.warning('failed to retrieve calculation results from xls w/ \n    %s'%e)
            calc_d = None
            
        #=======================================================================
        # get discounted values
        #=======================================================================
        ylab = self.impact_units
        legendTitle = 'Future Values'
        if presentVal:
            """should have worked w/ transposed...."""
            dfp = dfc.append(pd.Series(range(len(df.columns)), index=df.columns, name='period'))
            
            dfp = dfp.append(pd.Series(calc_d['discount_rate']+1, index=df.columns, name='rate'
                                       ).pow(dfp.loc['period', :]).rename('scale'))
            
            
            #get scaled values
            cols_d = dict()
            for indxr in df.index:
                cols_d[indxr] = '%s (PV)'%indxr
                dfp.loc[cols_d[indxr]] = dfp.loc[indxr, :].divide(dfp.loc['scale', :])
                
            #set new plot frame
            """better to keep the names the same here for plot styling"""
            dfc = dfp.loc[cols_d.values(), :].rename(index={v:k for k,v in cols_d.items()})
            
            #update lables
            ylab = ylab + ' (PV)'
            title = title + ' (PV)'
            legendTitle = 'Present Values'
        
        #=======================================================================
        # plot----
        #=======================================================================
        #======================================================================
        # figure setup
        #======================================================================
        """
        plt.show()
        """
        plt.close()
        fig = plt.figure(figsize=self.figsize, constrained_layout = True)
        
        #axis setup
        ax = fig.add_subplot(111)

        # axis label setup
        fig.suptitle(title)
        ax.set_ylabel(ylab)
        ax.set_xlabel('year')
        
        #=======================================================================
        # add lines
        #=======================================================================
        fill_d = dict()
        for serName, row in dfc.iterrows():
            #get style
            if serName in style_lib:
                styles = style_lib[serName]
            else:
                styles = dict()
            
            xar,  yar = row.index.values, row.values
            pline1 = ax.plot(xar,yar,
                            label       = serName.replace('Grand', '').strip(),
                            **styles
                            )
            
            #collect for filling
            for k in ['benefit', 'cost']:
                if k in serName.lower():
                    fill_d[k] = yar
                    break

            
        #=======================================================================
        # fill between the lines
        #=======================================================================
        if len(fill_d)==2:
            if 'hatch' in style_lib:
                styles = style_lib['hatch']
            else:
                styles = dict()
                
            for regionName, (pwhere, color) in {
                'Positive Investment':(fill_d['benefit']>fill_d['cost'], 'green'),
                'Negative Investment':(fill_d['benefit']<fill_d['cost'], 'red'),
                    }.items():
 
                if not np.any(pwhere): continue  #nothing here.. skip
                
                polys = ax.fill_between(xar.astype(np.float), fill_d['benefit'], 
                                        y2=fill_d['cost'],
                                        where=pwhere, color=color, label=regionName,
                                         **styles)
            
        #=======================================================================
        # post format------
        #=======================================================================
        #=======================================================================
        # text box
        #=======================================================================
        val_str = ''
        if not calc_d is None:
            for k,v in calc_d.items():
                val_str = val_str + '\n%s=%s'%(k,v)
        
        self._postFmt(ax, val_str=val_str, xLocScale=0.8, legendTitle=legendTitle)
        
        #assign tick formatter functions

            
        self._tickSet(ax, yfmtFunc=impactFmtFunc)
        
        log.info('finished')
        
        return fig
    
        """
        plt.show()
        view(df)
        view(df_raw)
        """
    
    def _get_cba_calcs(self, #helper to get teh calculation results from teh spreadsheet
                       data_fp=None,
                       logger=None,
                       section_nm = 'Cost-Benefit Calculations',
                       impactFmtFunc=None,
                       ):
        if logger is None: logger=self.logger
        log=logger.getChild('cba_calcs')
        
        if data_fp is None: data_fp = self.cba_xls
        assert os.path.exists(data_fp), 'passed bad filepath for cba_xls: %s'%data_fp
        
        df_raw = pd.read_excel(data_fp, sheet_name='smry' ,engine='openpyxl',
                               index_col=0)
        
        df = df_raw.loc[df_raw.index.notna(), :]
        df.index.name=None
        #=======================================================================
        # #get everything beneath Cost-Benefit Calculations
        #=======================================================================
        assert section_nm in df.index, 'unable to locate section: \'%s\''%section_nm

        d = df.iloc[df.index.get_loc(section_nm)+1:,0].to_dict()
        
        #=======================================================================
        # apply formatters
        #=======================================================================
        
        for k,v in d.copy().items():

            if k.endswith('$'):
                if not impactFmtFunc is None:
                    d[k] = impactFmtFunc(v)
                
            
            elif k.endswith('ratio'):
                d[k] = '%.4f'%v
                    
                
                    
                
        return d
        
        
        
        
        
        
        
        
        
        
        
        

            
 
        
        
        
        
        
        
        
        

    
    
    

    

            
        